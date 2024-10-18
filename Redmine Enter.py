from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
import openpyxl

# Define paths
excel_file_path = "C:/Selenium/Excel/1.xlsx"
gecko_driver_path = "C:/Selenium/Drivers/geckodriver.exe"

# Initialize the Service object for GeckoDriver
service = Service(gecko_driver_path)

# Initialize the Firefox driver with options to ignore SSL errors
options = webdriver.FirefoxOptions()
options.accept_insecure_certs = True

# Start the Firefox driver using the Service object
driver = webdriver.Firefox(service=service, options=options)

# Set up Selenium WebDriver
driver.implicitly_wait(10)  # Implicit wait for elements to load
driver.maximize_window()

# Navigate to the login page
driver.get("http://issuetracker.ntc-us.com/redmine/login?back_url=http%3A%2F%2Fissuetracker.ntc-us.com%2Fredmine%2Fprojects%2Ftd-support-tickets%2Fissues%2Fnew")
driver.find_element(By.ID, "username").send_keys("anilv")
driver.find_element(By.ID, "password").send_keys("abcd1234")
driver.find_element(By.ID, "login-submit").click()
time.sleep(2)

# Select the issue tracker from the dropdown
dropdown = Select(driver.find_element(By.ID, "issue_tracker_id"))
dropdown.select_by_visible_text("Support")

# Open Excel file
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active
row_count = sheet.max_row

# Loop through Excel rows
for row in range(1, row_count + 1):
    issue = sheet.cell(row=row, column=1).value
    subject = sheet.cell(row=row, column=2).value
    subject1 = sheet.cell(row=row, column=3).value
    subject2 = sheet.cell(row=row, column=4).value
    description = sheet.cell(row=row, column=5).value
    category = sheet.cell(row=row, column=6).value

    # Enter data into the web application
    driver.find_element(By.ID, "issue_custom_field_values_5").send_keys(issue)
    time.sleep(0)

    driver.find_element(By.ID, "issue_subject").send_keys(subject)
    driver.find_element(By.ID, "issue_subject").send_keys(subject1)
    driver.find_element(By.ID, "issue_subject").send_keys(subject2)
    time.sleep(0)

    driver.find_element(By.ID, "issue_description").send_keys(description)
    time.sleep(0)

    dropdown_assigned_to = Select(driver.find_element(By.ID, "issue_assigned_to_id"))
    dropdown_assigned_to.select_by_visible_text("Anil vasara")
    time.sleep(0)

    dropdown_category = Select(driver.find_element(By.ID, "issue_category_id"))
    dropdown_category.select_by_visible_text(category)
    time.sleep(2)

    driver.find_element(By.CSS_SELECTOR, ".list_cf > label:nth-child(1) > input:nth-child(1)").click()
    #driver.find_element(By.CSS_SELECTOR, ".list_cf > label:nth-child(2) > input:nth-child(1)").click()
    time.sleep(1)

    driver.find_element(By.CSS_SELECTOR, "#issue-form > input:nth-child(5)").click()
    time.sleep(1)

# Close Excel (not needed in Python as it doesn't keep an open instance)
workbook.close()

# Close the browser after completion
driver.quit()
